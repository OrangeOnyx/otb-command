# -*- coding: utf-8 -*-
"""OTB owner proforma — live Excel operating statement.
Income is real (units.json base + recoveries.json CAM/Tax/Ins, single-sourced per
the SOT rule). Operating expenses are seeded with defensible estimates the owner
overrides with actuals; NOI and valuation are live formulas. `npm run proforma`."""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
units = json.load(open(os.path.join(ROOT, "src", "data", "units.json"), encoding="utf-8"))
rec = json.load(open(os.path.join(ROOT, "src", "data", "recoveries.json"), encoding="utf-8"))["units"]
OUT = os.path.join(ROOT, "export"); os.makedirs(OUT, exist_ok=True)

# ---- order units by numeric unit ----
def ukey(u):
    s = u["unit"]; return float(s[:-1]) + 0.5 if s and s[-1].isalpha() else float(s)
units = sorted(units, key=ukey)

NAVY="1C2D4F"; INK="1C2B26"; PAPER="F5F5F5"; TAN="EDE7D6"; INPUT="FFF6D8"; GREEN="2F6B4F"
thin = Side(style="thin", color="C9C2AE")
def money(c): c.number_format = '#,##0;(#,##0)'
def psf(c): c.number_format = '#,##0.00'
def hdr(c): c.font=Font(bold=True,color="FFFFFF",size=10); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

wb = openpyxl.Workbook()

# ================= Rent Roll =================
rr = wb.active; rr.title = "Rent Roll"
rr["A1"]="On The Boulevard — Rent Roll (in-place income, from SOT)"; rr["A1"].font=Font(bold=True,size=13,color=NAVY)
cols=[("Unit",8),("Tenant",26),("SF",9),("Base PSF",9),("CAM PSF",9),("Tax PSF",9),("Ins PSF",9),
      ("Total PSF",9),("Base $",12),("CAM $",11),("Tax $",11),("Ins $",11),("Total $",13),("Status",10),("Term end",11)]
H=3
for i,(name,w) in enumerate(cols):
    c=rr.cell(H,i+1,name); hdr(c); rr.column_dimensions[get_column_letter(i+1)].width=w
r=H+1
for u in units:
    uid=u["unit"]; rr_= rec.get(uid,{})
    base=u.get("base",0) or 0; cam=rr_.get("cam",0); tax=rr_.get("tax",0); ins=rr_.get("ins",0); sf=u["sf"]
    rr.cell(r,1,uid); rr.cell(r,2,u["dba"]); rr.cell(r,3,sf)
    rr.cell(r,4,base); rr.cell(r,5,cam); rr.cell(r,6,tax); rr.cell(r,7,ins)
    rr.cell(r,8,f"=D{r}+E{r}+F{r}+G{r}")
    rr.cell(r,9,f"=C{r}*D{r}"); rr.cell(r,10,f"=C{r}*E{r}"); rr.cell(r,11,f"=C{r}*F{r}"); rr.cell(r,12,f"=C{r}*G{r}")
    rr.cell(r,13,f"=I{r}+J{r}+K{r}+L{r}")
    rr.cell(r,14,u["status"]); rr.cell(r,15,u.get("end","") or "")
    for col in (3,9,10,11,12,13): money(rr.cell(r,col))
    for col in (4,5,6,7,8): psf(rr.cell(r,col))
    r+=1
TOT=r
rr.cell(TOT,2,"TOTAL / occupied in-place").font=Font(bold=True)
rr.cell(TOT,3,f"=SUM(C{H+1}:C{TOT-1})")
for col in (9,10,11,12,13):
    L=get_column_letter(col); rr.cell(TOT,col,f"=SUM({L}{H+1}:{L}{TOT-1})")
for col in (3,9,10,11,12,13):
    c=rr.cell(TOT,col); c.font=Font(bold=True); money(c); c.border=Border(top=thin)
rr.freeze_panes="A4"
# named refs for the proforma
SF_T=f"'Rent Roll'!C{TOT}"; BASE_T=f"'Rent Roll'!I{TOT}"; CAM_T=f"'Rent Roll'!J{TOT}"; TAX_T=f"'Rent Roll'!K{TOT}"; INS_T=f"'Rent Roll'!L{TOT}"

# numeric prefills (estimates) computed here
sf_sum=sum(u["sf"] for u in units)
base_t=sum((u.get("base",0) or 0)*u["sf"] for u in units)
cam_t=sum(rec.get(u["unit"],{}).get("cam",0)*u["sf"] for u in units)
tax_t=sum(rec.get(u["unit"],{}).get("tax",0)*u["sf"] for u in units)
ins_t=sum(rec.get(u["unit"],{}).get("ins",0)*u["sf"] for u in units)
egi=base_t+cam_t+tax_t+ins_t

# ================= Proforma =================
pf = wb.create_sheet("Proforma")
pf.column_dimensions["A"].width=3
pf.column_dimensions["B"].width=34
pf.column_dimensions["C"].width=15
pf.column_dimensions["D"].width=11
pf.column_dimensions["E"].width=40
pf["B1"]="On The Boulevard Shopping Center"; pf["B1"].font=Font(bold=True,size=15,color=NAVY)
pf["B2"]="Annual Operating Proforma — 101–149 Arnould Blvd, Lafayette, LA 70506"; pf["B2"].font=Font(size=10,color=INK)
pf["B3"]="Owner: Belle Realty of Lafayette, LLC · mgr Orange Ocean, LLC · GLA 62,883 SF · 27 units"; pf["B3"].font=Font(size=9,italic=True,color="6b6450")

row=5
def section(t):
    global row
    c=pf.cell(row,2,t); c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY)
    for cc in range(2,6): pf.cell(row,cc).fill=PatternFill("solid",fgColor=NAVY)
    pf.cell(row,3,"$ / yr").font=Font(bold=True,color="FFFFFF"); pf.cell(row,4,"$/SF").font=Font(bold=True,color="FFFFFF")
    row+=1
def line(label, formula, note="", bold=False, input_=False, fill=None):
    global row
    pf.cell(row,2,label).font=Font(bold=bold)
    c=pf.cell(row,3,formula); money(c); c.font=Font(bold=bold)
    if input_: c.fill=PatternFill("solid",fgColor=INPUT)
    if fill:
        for cc in range(2,5): pf.cell(row,cc).fill=PatternFill("solid",fgColor=fill)
    d=pf.cell(row,4,f"=IF($C${SFcell}=0,0,C{row}/$C${SFcell})"); psf(d); d.font=Font(bold=bold,color="6b6450")
    if note: pf.cell(row,5,note).font=Font(size=9,italic=True,color="6b6450")
    cur=row; row+=1; return cur

# stash SF total on the sheet for $/SF math
pf.cell(row,2,"Gross leasable area (SF)").font=Font(bold=False)
SFcell=row
csf=pf.cell(row,3,sf_sum); csf.number_format='#,##0'
pf.cell(row,5,"Sum of demised SF (headline GLA 62,883)").font=Font(size=9,italic=True,color="6b6450")
row+=2

section("INCOME (in-place, annual)")
r_base=line("Base rent", f"={BASE_T}", "PSF × SF, single-sourced from units.json")
r_cam =line("CAM recovery (NNN)", f"={CAM_T}", "tenant reimbursement")
r_tax =line("Real-estate-tax recovery (NNN)", f"={TAX_T}", "tenant reimbursement")
r_ins =line("Insurance recovery (NNN)", f"={INS_T}", "tenant reimbursement")
r_egi =line("Effective Gross Income (EGI)", f"=C{r_base}+C{r_cam}+C{r_tax}+C{r_ins}", "100% in-place; 2 vacancies excluded (upside)", bold=True, fill=TAN)
row+=1

section("OPERATING EXPENSES (annual — replace estimates with actuals)")
r_etax=line("Real estate taxes", round(tax_t), "EST = tax recovery; recoverable (NNN)", input_=True)
r_eins=line("Insurance", round(ins_t), "EST = ins recovery; recoverable (NNN)", input_=True)
r_ecam=line("CAM / R&M (common area)", round(cam_t), "EST = CAM recovery; recoverable (NNN)", input_=True)
r_emgt=line("Management fee", round(0.04*egi), "EST 4% of EGI — confirm contract", input_=True)
r_eutl=line("Utilities (common / non-recoverable)", 0, "owner actual", input_=True)
r_ersv=line("Reserves / replacement", round(0.15*sf_sum), "EST $0.15/SF", input_=True)
r_opex=line("Total operating expenses", f"=SUM(C{r_etax}:C{r_ersv})", "", bold=True, fill=TAN)
row+=1

section("NET OPERATING INCOME")
r_noi=line("NOI (EGI − OpEx)", f"=C{r_egi}-C{r_opex}", "", bold=True, fill="DDE9DE")
r_net=line("Net NNN exposure", f"=(C{r_cam}+C{r_tax}+C{r_ins})-(C{r_etax}+C{r_eins}+C{r_ecam})",
           "recoveries − recoverable expenses (≈0 when pass-throughs match)")
row+=1

section("INDICATED VALUE (NOI ÷ cap rate)")
for cap in (7.0, 7.5, 8.0):
    pf.cell(row,2,f"Value @ {cap:.1f}% cap").font=Font(bold=(cap==7.5))
    cv=pf.cell(row,3,f"=C{r_noi}/{cap/100}"); money(cv); cv.font=Font(bold=(cap==7.5),color=GREEN)
    dv=pf.cell(row,4,f"=IF($C${SFcell}=0,0,C{row}/$C${SFcell})"); psf(dv)
    pf.cell(row,5,f"$/SF on {sf_sum:,} SF").font=Font(size=9,italic=True,color="6b6450")
    row+=1
row+=1

# notes
notes=[
 "METHODOLOGY & ASSUMPTIONS",
 "• Income is in-place and real: base rent per units.json (SOT), CAM/Tax/Ins per recoveries.json. EGI excludes the 2 vacancies (upside not shown).",
 "• Operating expenses are NOT in the SOT. Yellow cells are seeded ESTIMATES — replace with the owner's actual figures; NOI and value update automatically.",
 "• Recoverable lines (taxes, insurance, CAM) are seeded at the NNN recovery amount, so net exposure ≈ $0 when pass-throughs match actual spend. Management seeded at 4% of EGI; reserves at $0.15/SF; utilities at $0 (set actual common-area utilities).",
 "• Cap-rate values are illustrative; the 7.5% line is centered, not an appraisal.",
 "• Vacancy/credit-loss line not modeled (property is shown at 100% in-place); add a vacancy allowance if underwriting to stabilized.",
 "• Figures trace to OTB Property Command; regenerate with `npm run proforma`.",
]
for i,t in enumerate(notes):
    c=pf.cell(row+i,2,t); c.font=Font(bold=(i==0),size=9,color=(NAVY if i==0 else "6b6450")); c.alignment=Alignment(wrap_text=False)
pf.sheet_view.showGridLines=False

wb.save(os.path.join(OUT,"OTB-Proforma.xlsx"))
print("proforma ->", os.path.join(OUT,"OTB-Proforma.xlsx"))
print("EGI in-place ~ $%s/yr (base %s + NNN %s)"%(format(round(egi),","),format(round(base_t),","),format(round(cam_t+tax_t+ins_t),",")))
